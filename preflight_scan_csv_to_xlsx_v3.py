import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
import io

def convert_csv_xlsx_sort_and_format(input_file: str, output_file: str):
    # Read the CSV file using Pandas
    df = pd.read_csv(input_file)

    # Sort the DataFrame by status and test case name
    df = df.sort_values(by=['Status', 'Test Case'], key=lambda x: x.map({'FAILED': 0, 'NOT_APP': 1, 'PASSED': 2}))

    # Create an Excel workbook using openpyxl
    wb = Workbook()
    ws = wb.active

    # Set column alignment
    alignment = Alignment(horizontal='center', vertical='center')
    for col in ws.columns:
        for cell in col:
            cell.alignment = alignment

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20

    # Write the data to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Set text wrap for column C
    for cell in ws['C']:
        cell.alignment = Alignment(wrap_text=True)  # Enable text wrap for column C

    # Set the cell background color for the data rows
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):  # Assuming 'Status' column is in column D
        for cell in row:
            if cell.value == 'PASSED':
                cell.font = Font(color='006400')  # Dark green font for 'PASSED'
            elif cell.value == 'FAILED':
                cell.font = Font(color='FF0000')  # Red font for 'FAILED'
            elif cell.value == 'NOT_APP':
                cell.font = Font(color='FFA500')  # Dark orange font for 'NOT_APP' (FFA500 is the hexadecimal color for dark orange)

    # Set column alignment
    for col in ws.columns:
        if col[0].value == 'Status' or col[0].value == 'Image Tag':
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            for cell in col:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Set the cell background color and font for the header row
    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Save the workbook
    wb.save(output_file)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Sort and format a CSV file and save as Excel workbook')
    parser.add_argument('input_file', type=str, help='Path to input CSV file')
    parser.add_argument('output_file', type=str, help='Path to output Excel file')
    args = parser.parse_args()
    convert_csv_xlsx_sort_and_format(args.input_file, args.output_file)
