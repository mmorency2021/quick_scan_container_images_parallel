#!/usr/bin/env python3
"""
A Python script implementation of a container image scanning tool that supports
parallel preflight scans and writes results directly to an XLSX file and HTML report.
It supports both API‐based mode (using an API token) and offline mode (using an image list file).
"""

import argparse
import sys
import os
import subprocess
import re
import time
import json
import csv
import shutil
import datetime
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict, Any, Optional
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ------------------------------------------------------------------------------
# HTML Report Generation Function
# ------------------------------------------------------------------------------

def write_html_report(scan_data: List[List[str]], detailed_checks: List[Dict[str, Any]], output_html: str) -> None:
    """
    Generate a comprehensive HTML report from scan results and detailed check information.
    """
    if not scan_data:
        raise ValueError("No data to write to HTML report")
    
    # Convert scan data to dictionary format for easier processing
    headers = ["Image Name", "Image Tag", "Has Modified Files", "Test Case", "Status"]
    dict_data = []
    for row in scan_data:
        if len(row) >= 5:
            dict_data.append({
                headers[0]: row[0],
                headers[1]: row[1], 
                headers[2]: row[2],
                headers[3]: row[3],
                headers[4]: row[4]
            })
    
    # Analyze data for statistics
    total_tests = len(dict_data)
    unique_images = set(row[0] for row in scan_data)
    total_unique_images = len(unique_images)
    
    # Count test results
    status_counts = Counter(row['Status'] for row in dict_data)
    passed_count = status_counts.get('PASSED', 0)
    failed_count = status_counts.get('FAILED', 0)
    not_app_count = status_counts.get('NOT_APP', 0)
    
    # Calculate percentages
    passed_pct = (passed_count / total_tests * 100) if total_tests > 0 else 0
    failed_pct = (failed_count / total_tests * 100) if total_tests > 0 else 0
    not_app_pct = (not_app_count / total_tests * 100) if total_tests > 0 else 0
    
    # Analyze by test case
    test_case_stats = defaultdict(lambda: {'total': 0, 'passed': 0, 'failed': 0, 'not_app': 0})
    for row in dict_data:
        test_case = row['Test Case']
        test_case_stats[test_case]['total'] += 1
        if row['Status'] == 'PASSED':
            test_case_stats[test_case]['passed'] += 1
        elif row['Status'] == 'FAILED':
            test_case_stats[test_case]['failed'] += 1
        else:
            test_case_stats[test_case]['not_app'] += 1
    
    # Find failed images and group failures
    failed_images = defaultdict(list)
    failed_test_details = defaultdict(list)
    
    for row in dict_data:
        if row['Status'] == 'FAILED':
            failed_images[row['Image Name']].append(row['Test Case'])
            failed_test_details[row['Test Case']].append(row['Image Name'])
    
    # Adaptive image categorization approach
    def categorize_images_adaptively(images):
        """Automatically detect common patterns and create categories."""
        from collections import Counter
        import re
        
        # Extract potential prefixes and patterns
        prefixes = []
        for img in images:
            # Try different separation patterns
            parts = re.split(r'[-_/]', img.lower())
            if len(parts) >= 2:
                prefixes.append(f"{parts[0]}-{parts[1]}")  # First two parts
            elif len(parts) >= 1:
                prefixes.append(parts[0])  # Just first part
        
        # Count prefix occurrences
        prefix_counts = Counter(prefixes)
        
        # Only create categories for prefixes that appear 2+ times
        categories = {}
        categorized_images = set()
        
        for prefix, count in prefix_counts.most_common():
            if count >= 2:  # At least 2 images with this pattern
                matching_images = [img for img in images if prefix.lower() in img.lower()]
                if len(matching_images) >= 2:
                    categories[f"{prefix}* Images"] = matching_images
                    categorized_images.update(matching_images)
        
        # Add uncategorized images
        uncategorized = [img for img in images if img not in categorized_images]
        if uncategorized:
            categories["Other Images"] = uncategorized
            
        return categories
    
    # Option 1: Use adaptive categorization (comment out to disable)
    adaptive_categories = categorize_images_adaptively(list(unique_images))
    
    # Option 2: Simple registry-based categorization (comment out to disable)  
    # registry_categories = {}
    # for img in unique_images:
    #     registry = img.split('/')[0] if '/' in img else 'local'
    #     if registry not in registry_categories:
    #         registry_categories[registry] = []
    #     registry_categories[registry].append(img)
    
    # Option 3: No categorization - just use perfect score vs issues (uncomment to use)
    # adaptive_categories = None
    
    # Count perfect score vs issues
    images_with_issues = len(failed_images)
    images_perfect = total_unique_images - images_with_issues
    
    # Get current timestamp
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    report_date = datetime.datetime.now().strftime("%B %d, %Y")
    
    # Generate HTML content
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Container Image Security Scanning Report</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
    <style>
        :root {{
            --primary-color: #667eea;
            --primary-dark: #564ba2;
            --success-color: #10b981;
            --success-dark: #059669;
            --error-color: #ef4444;
            --error-dark: #dc2626;
            --warning-color: #f59e0b;
            --warning-dark: #d97706;
            --gray-50: #f9fafb;
            --gray-100: #f3f4f6;
            --gray-200: #e5e7eb;
            --gray-300: #d1d5db;
            --gray-400: #9ca3af;
            --gray-500: #6b7280;
            --gray-600: #4b5563;
            --gray-700: #374151;
            --gray-800: #1f2937;
            --gray-900: #111827;
            --shadow-sm: 0 1px 2px 0 rgb(0 0 0 / 0.05);
            --shadow: 0 1px 3px 0 rgb(0 0 0 / 0.1), 0 1px 2px -1px rgb(0 0 0 / 0.1);
            --shadow-md: 0 4px 6px -1px rgb(0 0 0 / 0.1), 0 2px 4px -2px rgb(0 0 0 / 0.1);
            --shadow-lg: 0 10px 15px -3px rgb(0 0 0 / 0.1), 0 4px 6px -4px rgb(0 0 0 / 0.1);
            --shadow-xl: 0 20px 25px -5px rgb(0 0 0 / 0.1), 0 8px 10px -6px rgb(0 0 0 / 0.1);
        }}

        * {{
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Inter', 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, var(--gray-50) 0%, var(--gray-100) 100%);
            color: var(--gray-800);
            line-height: 1.6;
            min-height: 100vh;
        }}

        /* Loading Animation */
        .loading-spinner {{
            display: none;
            position: fixed;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            z-index: 10000;
        }}

        .spinner {{
            width: 40px;
            height: 40px;
            border: 4px solid var(--gray-200);
            border-top: 4px solid var(--primary-color);
            border-radius: 50%;
            animation: spin 1s linear infinite;
        }}

        @keyframes spin {{
            0% {{ transform: rotate(0deg); }}
            100% {{ transform: rotate(360deg); }}
        }}

        /* Tooltip Styles */
        .tooltip {{
            position: relative;
            cursor: help;
        }}

        .tooltip::before {{
            content: attr(data-tooltip);
            position: absolute;
            bottom: 125%;
            left: 50%;
            transform: translateX(-50%);
            background-color: var(--gray-900);
            color: white;
            padding: 8px 12px;
            border-radius: 6px;
            font-size: 0.875rem;
            font-weight: 400;
            white-space: nowrap;
            z-index: 1000;
            opacity: 0;
            visibility: hidden;
            transition: all 0.3s ease;
            box-shadow: var(--shadow-lg);
        }}

        .tooltip::after {{
            content: '';
            position: absolute;
            bottom: 115%;
            left: 50%;
            transform: translateX(-50%);
            border: 5px solid transparent;
            border-top-color: var(--gray-900);
            z-index: 1000;
            opacity: 0;
            visibility: hidden;
            transition: all 0.3s ease;
        }}

        .tooltip:hover::before,
        .tooltip:hover::after {{
            opacity: 1;
            visibility: visible;
        }}

        /* Copy Button */
        .copy-btn {{
            background: var(--gray-100);
            border: 1px solid var(--gray-300);
            border-radius: 6px;
            padding: 4px 8px;
            font-size: 0.75rem;
            cursor: pointer;
            transition: all 0.2s ease;
            margin-left: 8px;
        }}

        .copy-btn:hover {{
            background: var(--gray-200);
            transform: translateY(-1px);
        }}

        .copy-btn.copied {{
            background: var(--success-color);
            color: white;
            border-color: var(--success-dark);
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background-color: white;
            box-shadow: var(--shadow-xl);
            border-radius: 16px;
            overflow: hidden;
            position: relative;
        }}
        
        /* Action Bar */
        .action-bar {{
            position: sticky;
            top: 20px;
            z-index: 100;
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(10px);
            border: 1px solid var(--gray-200);
            border-radius: 12px;
            padding: 12px 20px;
            margin-bottom: 20px;
            display: flex;
            align-items: center;
            gap: 16px;
            flex-wrap: wrap;
            box-shadow: var(--shadow);
        }}

        .search-box {{
            position: relative;
            flex: 1;
            min-width: 250px;
        }}

        .search-box input {{
            width: 100%;
            padding: 8px 12px 8px 36px;
            border: 1px solid var(--gray-300);
            border-radius: 8px;
            font-size: 0.9rem;
            transition: all 0.2s ease;
        }}

        .search-box input:focus {{
            outline: none;
            border-color: var(--primary-color);
            box-shadow: 0 0 0 3px rgb(102 126 234 / 0.1);
        }}

        .search-box i {{
            position: absolute;
            left: 12px;
            top: 50%;
            transform: translateY(-50%);
            color: var(--gray-400);
        }}

        .action-buttons {{
            display: flex;
            gap: 8px;
            align-items: center;
        }}

        .btn {{
            padding: 8px 16px;
            border: 1px solid var(--gray-300);
            border-radius: 8px;
            background: white;
            color: var(--gray-700);
            font-size: 0.875rem;
            font-weight: 500;
            cursor: pointer;
            transition: all 0.2s ease;
            display: inline-flex;
            align-items: center;
            gap: 6px;
            text-decoration: none;
        }}

        .btn:hover {{
            background: var(--gray-50);
            transform: translateY(-1px);
            box-shadow: var(--shadow-sm);
        }}

        .btn-primary {{
            background: var(--primary-color);
            color: white;
            border-color: var(--primary-dark);
        }}

        .btn-primary:hover {{
            background: var(--primary-dark);
        }}
        
        .header {{
            background: linear-gradient(135deg, var(--primary-color) 0%, var(--primary-dark) 100%);
            color: white;
            padding: 40px 40px 30px;
            text-align: center;
            position: relative;
            overflow: hidden;
        }}

        .header::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: url('data:image/svg+xml,<svg width="60" height="60" viewBox="0 0 60 60" xmlns="http://www.w3.org/2000/svg"><g fill="none" fill-rule="evenodd"><g fill="%23ffffff" fill-opacity="0.05"><circle cx="7" cy="7" r="4"/><circle cx="53" cy="53" r="4"/><circle cx="53" cy="7" r="4"/><circle cx="7" cy="53" r="4"/></g></g></svg>');
            opacity: 0.3;
        }}
        
        .header h1 {{
            margin: 0;
            font-size: 2.75rem;
            font-weight: 700;
            position: relative;
            z-index: 1;
        }}
        
        .header p {{
            margin: 12px 0 0 0;
            font-size: 1.25rem;
            opacity: 0.9;
            font-weight: 400;
            position: relative;
            z-index: 1;
        }}

        .header-stats {{
            display: flex;
            justify-content: center;
            gap: 32px;
            margin-top: 24px;
            position: relative;
            z-index: 1;
        }}

        .header-stat {{
            text-align: center;
        }}

        .header-stat-number {{
            display: block;
            font-size: 1.875rem;
            font-weight: 700;
            margin-bottom: 4px;
        }}

        .header-stat-label {{
            display: block;
            font-size: 0.875rem;
            opacity: 0.8;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        
        .content {{
            padding: 40px;
        }}
        
        .section {{
            margin-bottom: 48px;
            opacity: 0;
            animation: fadeInUp 0.6s ease forwards;
        }}

        .section:nth-child(1) {{ animation-delay: 0.1s; }}
        .section:nth-child(2) {{ animation-delay: 0.2s; }}
        .section:nth-child(3) {{ animation-delay: 0.3s; }}
        .section:nth-child(4) {{ animation-delay: 0.4s; }}

        @keyframes fadeInUp {{
            from {{
                opacity: 0;
                transform: translateY(30px);
            }}
            to {{
                opacity: 1;
                transform: translateY(0);
            }}
        }}
        
        .section h2 {{
            color: var(--gray-900);
            font-size: 1.875rem;
            font-weight: 700;
            margin: 0 0 24px 0;
            display: flex;
            align-items: center;
            gap: 12px;
        }}

        .section h2::after {{
            content: '';
            flex: 1;
            height: 2px;
            background: linear-gradient(90deg, var(--primary-color), transparent);
        }}

        .section-icon {{
            font-size: 1.5rem;
            color: var(--primary-color);
        }}
        
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 24px;
            margin-bottom: 32px;
        }}
        
        .stat-card {{
            background: white;
            border: 1px solid var(--gray-200);
            border-radius: 12px;
            padding: 32px 24px;
            text-align: center;
            box-shadow: var(--shadow);
            transition: all 0.3s ease;
            position: relative;
            overflow: hidden;
            text-decoration: none;
            color: inherit;
            display: block;
        }}

        .stat-card.clickable {{
            cursor: pointer;
        }}

        .stat-card.clickable:hover {{
            transform: translateY(-6px);
            box-shadow: var(--shadow-xl);
            border-color: var(--primary-color);
        }}

        .stat-card.clickable::after {{
            content: '🔗';
            position: absolute;
            top: 12px;
            right: 12px;
            font-size: 1rem;
            opacity: 0;
            transition: opacity 0.3s ease;
        }}

        .stat-card.clickable:hover::after {{
            opacity: 0.6;
        }}

        .stat-card::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
            background: var(--gray-300);
            transition: all 0.3s ease;
        }}
        
        .stat-card.passed::before {{
            background: linear-gradient(90deg, var(--success-color), var(--success-dark));
        }}
        
        .stat-card.failed::before {{
            background: linear-gradient(90deg, var(--error-color), var(--error-dark));
        }}
        
        .stat-card.not-app::before {{
            background: linear-gradient(90deg, var(--warning-color), var(--warning-dark));
        }}

        .stat-card:hover:not(.clickable) {{
            transform: translateY(-4px);
            box-shadow: var(--shadow-lg);
        }}
        
        .stat-icon {{
            font-size: 2.5rem;
            margin-bottom: 16px;
            opacity: 0.8;
        }}

        .stat-card.passed .stat-icon {{ color: var(--success-color); }}
        .stat-card.failed .stat-icon {{ color: var(--error-color); }}
        .stat-card.not-app .stat-icon {{ color: var(--warning-color); }}
        .stat-card .stat-icon {{ color: var(--primary-color); }}
        
        .stat-number {{
            font-size: 3rem;
            font-weight: 700;
            margin-bottom: 8px;
            color: var(--gray-900);
            line-height: 1;
        }}
        
        .stat-label {{
            font-size: 1rem;
            font-weight: 500;
            color: var(--gray-600);
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .stat-sublabel {{
            font-size: 0.875rem;
            color: var(--gray-500);
            margin-top: 4px;
            font-weight: 400;
            text-transform: none;
            letter-spacing: normal;
        }}
        
        .table-container {{
            margin-top: 24px;
            border: 1px solid var(--gray-200);
            border-radius: 12px;
            overflow: hidden;
            box-shadow: var(--shadow);
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            background: white;
        }}
        
        th {{
            background: var(--gray-50);
            color: var(--gray-900);
            padding: 16px 20px;
            text-align: left;
            font-weight: 600;
            font-size: 0.875rem;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            border-bottom: 1px solid var(--gray-200);
        }}
        
        td {{
            padding: 16px 20px;
            border-bottom: 1px solid var(--gray-100);
            font-size: 0.9rem;
            vertical-align: top;
        }}
        
        tr:last-child td {{
            border-bottom: none;
        }}
        
        tr:hover {{
            background-color: var(--gray-50);
        }}

        .searchable-table tr {{
            transition: opacity 0.2s ease;
        }}

        .searchable-table tr.hidden {{
            display: none;
        }}
        
        .status-badge {{
            display: inline-flex;
            align-items: center;
            gap: 4px;
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 0.75rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            border: 1px solid;
        }}
        
        .status-passed {{
            background-color: rgb(16 185 129 / 0.1);
            color: var(--success-dark);
            border-color: rgb(16 185 129 / 0.2);
        }}
        
        .status-failed {{
            background-color: rgb(239 68 68 / 0.1);
            color: var(--error-dark);
            border-color: rgb(239 68 68 / 0.2);
        }}
        
        .status-not-app {{
            background-color: rgb(245 158 11 / 0.1);
            color: var(--warning-dark);
            border-color: rgb(245 158 11 / 0.2);
        }}

        /* Enhanced tooltips for status badges */
        .status-badge.tooltip[data-tooltip*="UBI"]::before {{
            width: 300px;
            white-space: normal;
            text-align: center;
        }}

        /* Enhanced Test Cases Analysis Table */
        .test-analysis-table {{
            background: white;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: var(--shadow-lg);
        }}

        .test-analysis-table th {{
            background: linear-gradient(135deg, var(--primary-color), var(--primary-dark));
            color: white;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            padding: 20px 16px;
            position: relative;
        }}

        .test-analysis-table th::after {{
            content: '';
            position: absolute;
            bottom: 0;
            left: 0;
            right: 0;
            height: 2px;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.3), transparent);
        }}

        .test-analysis-table td {{
            padding: 20px 16px;
            vertical-align: middle;
        }}

        .test-case-name {{
            font-weight: 600;
            color: var(--gray-800);
            position: relative;
        }}

        .test-case-icon {{
            margin-right: 8px;
            color: var(--primary-color);
            font-size: 1.1rem;
        }}

        .metric-number {{
            font-weight: 700;
            font-size: 1.1rem;
            padding: 6px 12px;
            border-radius: 6px;
            display: inline-block;
            min-width: 40px;
            text-align: center;
            transition: all 0.3s ease;
            opacity: 0;
            animation: slideInNumber 0.6s ease forwards;
        }}

        @keyframes slideInNumber {{
            from {{
                opacity: 0;
                transform: translateY(10px) scale(0.8);
            }}
            to {{
                opacity: 1;
                transform: translateY(0) scale(1);
            }}
        }}

        .test-analysis-table tbody tr:nth-child(1) .metric-number {{ animation-delay: 0.1s; }}
        .test-analysis-table tbody tr:nth-child(2) .metric-number {{ animation-delay: 0.2s; }}
        .test-analysis-table tbody tr:nth-child(3) .metric-number {{ animation-delay: 0.3s; }}
        .test-analysis-table tbody tr:nth-child(4) .metric-number {{ animation-delay: 0.4s; }}
        .test-analysis-table tbody tr:nth-child(5) .metric-number {{ animation-delay: 0.5s; }}
        .test-analysis-table tbody tr:nth-child(6) .metric-number {{ animation-delay: 0.6s; }}
        .test-analysis-table tbody tr:nth-child(7) .metric-number {{ animation-delay: 0.7s; }}
        .test-analysis-table tbody tr:nth-child(8) .metric-number {{ animation-delay: 0.8s; }}
        .test-analysis-table tbody tr:nth-child(9) .metric-number {{ animation-delay: 0.9s; }}
        .test-analysis-table tbody tr:nth-child(10) .metric-number {{ animation-delay: 1.0s; }}

        .metric-number:hover {{
            transform: scale(1.1);
            box-shadow: var(--shadow-md);
        }}

        .metric-passed {{
            background: linear-gradient(135deg, rgb(16 185 129 / 0.15), rgb(16 185 129 / 0.1));
            color: var(--success-dark);
            border: 1px solid rgb(16 185 129 / 0.3);
        }}

        .metric-failed {{
            background: linear-gradient(135deg, rgb(239 68 68 / 0.15), rgb(239 68 68 / 0.1));
            color: var(--error-dark);
            border: 1px solid rgb(239 68 68 / 0.3);
        }}

        .metric-not-app {{
            background: linear-gradient(135deg, rgb(245 158 11 / 0.15), rgb(245 158 11 / 0.1));
            color: var(--warning-dark);
            border: 1px solid rgb(245 158 11 / 0.3);
        }}

        .metric-total {{
            background: linear-gradient(135deg, var(--gray-100), var(--gray-50));
            color: var(--gray-800);
            border: 1px solid var(--gray-300);
        }}

        .success-rate-container {{
            position: relative;
            padding: 4px 0;
        }}

        .success-rate-bar {{
            background: var(--gray-200);
            height: 8px;
            border-radius: 4px;
            overflow: hidden;
            margin-bottom: 4px;
        }}

        .success-rate-fill {{
            height: 100%;
            border-radius: 4px;
            transition: width 0.6s ease;
        }}

        .success-rate-fill.excellent {{
            background: linear-gradient(90deg, var(--success-color), var(--success-dark));
        }}

        .success-rate-fill.good {{
            background: linear-gradient(90deg, #22c55e, #16a34a);
        }}

        .success-rate-fill.warning {{
            background: linear-gradient(90deg, var(--warning-color), var(--warning-dark));
        }}

        .success-rate-fill.poor {{
            background: linear-gradient(90deg, var(--error-color), var(--error-dark));
        }}

        .success-rate-text {{
            font-weight: 600;
            font-size: 0.9rem;
            text-align: center;
        }}

        .success-rate-text.excellent {{
            color: var(--success-dark);
        }}

        .success-rate-text.good {{
            color: #16a34a;
        }}

        .success-rate-text.warning {{
            color: var(--warning-dark);
        }}

        .success-rate-text.poor {{
            color: var(--error-dark);
        }}

        .test-analysis-table tr:hover {{
            background: linear-gradient(135deg, var(--gray-50), rgba(102, 126, 234, 0.02));
            transform: scale(1.01);
            box-shadow: var(--shadow-sm);
        }}

        .test-analysis-table tbody tr {{
            transition: all 0.3s ease;
            border-left: 4px solid transparent;
        }}

        .test-analysis-table tbody tr:hover {{
            border-left-color: var(--primary-color);
        }}

        .metric-zero {{
            opacity: 0.5;
            font-style: italic;
        }}
        
        .chart-container {{
            margin: 20px 0;
            text-align: center;
        }}
        
        .progress-bar {{
            background-color: #e0e0e0;
            border-radius: 10px;
            overflow: hidden;
            height: 30px;
            margin: 10px 0;
        }}
        
        .progress-fill {{
            height: 100%;
            transition: width 0.3s ease;
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: bold;
        }}
        
        .progress-passed {{
            background: linear-gradient(90deg, #4CAF50, #45a049);
        }}
        
        .progress-failed {{
            background: linear-gradient(90deg, #f44336, #da190b);
        }}
        
        .progress-not-app {{
            background: linear-gradient(90deg, #ff9800, #e68900);
        }}
        
        .summary-box {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 25px;
            border-radius: 10px;
            margin: 20px 0;
        }}
        
        .timestamp {{
            color: #6c757d;
            font-size: 0.9em;
            text-align: center;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #dee2e6;
        }}
        
        .image-list {{
            background-color: #f8f9fa;
            padding: 20px;
            border-radius: 10px;
            margin: 20px 0;
        }}
        
        .image-item {{
            background-color: white;
            margin: 10px 0;
            padding: 15px;
            border-radius: 5px;
            border-left: 4px solid #667eea;
        }}
        
        .image-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }}
        
        .image-card {{
            background-color: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 8px;
            padding: 15px;
            transition: all 0.3s ease;
        }}
        
        .image-card:hover {{
            background-color: #e3f2fd;
            border-color: #667eea;
        }}
        
        .image-card h4 {{
            margin: 0 0 10px 0;
            color: #2c3e50;
            font-size: 1.1em;
        }}
        
        .status-summary {{
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }}
        
        .status-mini {{
            padding: 2px 8px;
            border-radius: 12px;
            font-size: 0.8em;
            font-weight: bold;
        }}
        
        /* Responsive Design */
        @media (max-width: 1024px) {{
            .container {{
                margin: 10px;
                border-radius: 12px;
            }}
            
            .content {{
                padding: 24px;
            }}
            
            .header-stats {{
                gap: 20px;
            }}
        }}

        @media (max-width: 768px) {{
            body {{
                padding: 10px;
            }}
            
            .stats-grid {{
                grid-template-columns: 1fr;
                gap: 16px;
            }}
            
            .header {{
                padding: 24px 20px;
            }}
            
            .header h1 {{
                font-size: 2rem;
            }}
            
            .header p {{
                font-size: 1rem;
            }}
            
            .header-stats {{
                flex-direction: column;
                gap: 16px;
            }}
            
            .content {{
                padding: 20px;
            }}
            
            .section h2 {{
                font-size: 1.5rem;
            }}
            
            .action-bar {{
                flex-direction: column;
                align-items: stretch;
            }}
            
            .search-box {{
                min-width: auto;
            }}
            
            .action-buttons {{
                justify-content: center;
            }}
            
            .table-container {{
                font-size: 0.875rem;
            }}
            
            th, td {{
                padding: 12px 16px;
            }}
        }}

        @media (max-width: 480px) {{
            .header h1 {{
                font-size: 1.75rem;
            }}
            
            .stat-number {{
                font-size: 2.5rem;
            }}
            
            .section h2 {{
                font-size: 1.25rem;
            }}
            
            th, td {{
                padding: 8px 12px;
                font-size: 0.8rem;
            }}
        }}

        /* Print Styles */
        @media print {{
            body {{
                background: white;
                padding: 0;
            }}
            
            .container {{
                box-shadow: none;
                border-radius: 0;
                margin: 0;
            }}
            
            .action-bar {{
                display: none;
            }}
            
            .header {{
                background: var(--gray-100) !important;
                color: var(--gray-900) !important;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            .stat-card {{
                border: 1px solid var(--gray-300);
                box-shadow: none;
                break-inside: avoid;
            }}
            
            .section {{
                break-inside: avoid;
            }}
            
            .copy-btn {{
                display: none;
            }}
            
            .tooltip::before,
            .tooltip::after {{
                display: none;
            }}
        }}

        /* Dark Mode Support */
        @media (prefers-color-scheme: dark) {{
            :root {{
                --gray-50: #18181b;
                --gray-100: #27272a;
                --gray-200: #3f3f46;
                --gray-300: #52525b;
                --gray-900: #f4f4f5;
                --gray-800: #e4e4e7;
            }}
        }}
    </style>
</head>
<body>
    <!-- Loading Spinner -->
    <div class="loading-spinner" id="loadingSpinner">
        <div class="spinner"></div>
    </div>

    <div class="container">
        <div class="header">
            <h1><i class="fas fa-shield-alt"></i> Container Image Security Scanning Report</h1>
            <p>Comprehensive Security and Compliance Analysis</p>
            <div class="header-stats">
                <div class="header-stat">
                    <span class="header-stat-number">{total_unique_images}</span>
                    <span class="header-stat-label">Images Scanned</span>
                </div>
                <div class="header-stat">
                    <span class="header-stat-number">{total_tests}</span>
                    <span class="header-stat-label">Total Tests</span>
                </div>
                <div class="header-stat">
                    <span class="header-stat-number">{passed_pct:.1f}%</span>
                    <span class="header-stat-label">Success Rate</span>
                </div>
            </div>
        </div>
        
        <!-- Action Bar -->
        <div class="action-bar">
            <div class="search-box">
                <i class="fas fa-search"></i>
                <input type="text" id="searchInput" placeholder="Search images, test cases, or status...">
            </div>
            <div class="action-buttons">
                <button class="btn" onclick="exportToCSV()">
                    <i class="fas fa-download"></i> Export CSV
                </button>
                <button class="btn" onclick="printReport()">
                    <i class="fas fa-print"></i> Print
                </button>
                <button class="btn btn-primary" onclick="copyReportLink()">
                    <i class="fas fa-link"></i> Share
                </button>
            </div>
        </div>
        
        <div class="content">
            <!-- Executive Summary -->
            <div class="section">
                <h2>📊 Executive Summary</h2>
                <div class="summary-box">
                    <p><strong>Scan Date:</strong> {report_date}</p>
                    <p><strong>Total Test Results:</strong> {total_tests} tests across {total_unique_images} unique container images</p>
                    <p><strong>Overall Status:</strong> {passed_count} Passed, {failed_count} Failed, {not_app_count} Not Applicable</p>
                    <p><strong>Success Rate:</strong> {passed_pct:.1f}% of tests passed successfully</p>
                </div>
            </div>
            
            <!-- Key Metrics -->
            <div class="section">
                <h2><i class="section-icon fas fa-chart-line"></i>Key Metrics</h2>
                <div class="stats-grid">
                    <a href="#detailed-test-results" class="stat-card passed clickable" onclick="filterTableByStatus('passed'); return true;">
                        <div class="stat-icon"><i class="fas fa-check-circle"></i></div>
                        <div class="stat-number">{passed_count}</div>
                        <div class="stat-label">Passed Tests</div>
                        <div class="stat-sublabel">{passed_pct:.1f}% Success Rate</div>
                    </a>"""

    # Only show failed tests link if there are failures
    if failed_count > 0:
        html_content += f"""
                    <a href="#failed-tests-details" class="stat-card failed clickable">
                        <div class="stat-icon"><i class="fas fa-exclamation-triangle"></i></div>
                        <div class="stat-number">{failed_count}</div>
                        <div class="stat-label">Failed Tests</div>
                        <div class="stat-sublabel">Requires Attention</div>
                    </a>"""
    else:
        html_content += f"""
                    <div class="stat-card failed">
                        <div class="stat-icon"><i class="fas fa-check-circle"></i></div>
                        <div class="stat-number">{failed_count}</div>
                        <div class="stat-label">No Failed Tests</div>
                        <div class="stat-sublabel">All Tests Passed!</div>
                    </div>"""

    # Only show NOT_APP link if there are any
    if not_app_count > 0:
        html_content += f"""
                    <a href="#detailed-test-results" class="stat-card not-app clickable" onclick="filterTableByStatus('not_app'); return true;">
                        <div class="stat-icon"><i class="fas fa-info-circle"></i></div>
                        <div class="stat-number">{not_app_count}</div>
                        <div class="stat-label">Not Applicable</div>
                        <div class="stat-sublabel">Non-UBI Based Images</div>
                    </a>"""
    else:
        html_content += f"""
                    <div class="stat-card not-app">
                        <div class="stat-icon"><i class="fas fa-check-circle"></i></div>
                        <div class="stat-number">{not_app_count}</div>
                        <div class="stat-label">All Tests Applicable</div>
                        <div class="stat-sublabel">UBI-Based Images</div>
                    </div>"""

    html_content += f"""
                    <a href="#detailed-test-results" class="stat-card clickable">
                        <div class="stat-icon"><i class="fas fa-cube"></i></div>
                        <div class="stat-number">{total_unique_images}</div>
                        <div class="stat-label">Unique Images</div>
                        <div class="stat-sublabel">Containers Analyzed</div>
                    </a>
                </div>
                
                <!-- Overall Progress -->
                <div class="chart-container">
                    <h3>Overall Test Results Distribution</h3>
                    <div class="progress-bar" style="display: flex; background-color: #e0e0e0;">"""

    # Create segments that add up to 100%
    segments = []
    
    if passed_count > 0:
        # Only show percentage text if segment is large enough
        text = f"{passed_pct:.1f}%" if passed_pct >= 10 else ""
        segments.append(f'<div class="progress-passed" style="width: {passed_pct:.1f}%; display: flex; align-items: center; justify-content: center; color: white; font-weight: bold; font-size: 0.9em;">{text}</div>')
    
    if failed_count > 0:
        text = f"{failed_pct:.1f}%" if failed_pct >= 10 else ""
        segments.append(f'<div class="progress-failed" style="width: {failed_pct:.1f}%; display: flex; align-items: center; justify-content: center; color: white; font-weight: bold; font-size: 0.9em;">{text}</div>')
    
    if not_app_count > 0:
        text = f"{not_app_pct:.1f}%" if not_app_pct >= 10 else ""
        segments.append(f'<div class="progress-not-app" style="width: {not_app_pct:.1f}%; display: flex; align-items: center; justify-content: center; color: white; font-weight: bold; font-size: 0.9em;">{text}</div>')

    html_content += ''.join(segments)
    
    html_content += f"""
                    </div>
                    <!-- Legend -->
                    <div style="margin-top: 15px; display: flex; justify-content: center; gap: 20px; flex-wrap: wrap; font-size: 0.95em;">
                        <div style="display: flex; align-items: center; gap: 5px;">
                            <div style="width: 16px; height: 16px; background: linear-gradient(90deg, #4CAF50, #45a049); border-radius: 3px;"></div>
                            <span><strong>{passed_pct:.1f}%</strong> Passed ({passed_count})</span>
                        </div>
                        <div style="display: flex; align-items: center; gap: 5px;">
                            <div style="width: 16px; height: 16px; background: linear-gradient(90deg, #f44336, #da190b); border-radius: 3px;"></div>
                            <span><strong>{failed_pct:.1f}%</strong> Failed ({failed_count})</span>
                        </div>"""
    
    if not_app_count > 0:
        html_content += f"""
                        <div style="display: flex; align-items: center; gap: 5px;">
                            <div style="width: 16px; height: 16px; background: linear-gradient(90deg, #ff9800, #e68900); border-radius: 3px;"></div>
                            <span><strong>{not_app_pct:.1f}%</strong> N/A ({not_app_count})</span>
                        </div>"""
    
    html_content += f"""
                    </div>
                    <div style="margin-top: 10px; text-align: center; font-size: 0.9em; color: #666;">
                        📊 Total: <strong>{total_tests}</strong> tests across <strong>{total_unique_images}</strong> images
                    </div>"""

    # Calculate summary statistics for Test Cases Analysis
    perfect_tests_count = len([tc for tc, stats in test_case_stats.items() if (stats['passed'] / stats['total'] * 100) == 100])
    total_test_executions = sum(stats['total'] for stats in test_case_stats.values())
    tests_with_failures = len([tc for tc, stats in test_case_stats.items() if stats['failed'] > 0])
    tests_with_not_app = len([tc for tc, stats in test_case_stats.items() if stats['not_app'] > 0])

    html_content += f"""
                </div>
            </div>
            
            <!-- Test Cases Analysis -->
            <div class="section" id="test-cases-analysis">
                <h2>🔍 Test Cases Analysis</h2>
                
                <!-- Quick Summary Cards -->
                <div class="stats-grid" style="margin-bottom: 32px;">
                    <div class="stat-card">
                        <div class="stat-icon"><i class="fas fa-clipboard-check"></i></div>
                        <div class="stat-number">{perfect_tests_count}</div>
                        <div class="stat-label">Perfect Tests</div>
                        <div class="stat-sublabel">100% Success Rate</div>
                    </div>
                    <div class="stat-card passed">
                        <div class="stat-icon"><i class="fas fa-chart-line"></i></div>
                        <div class="stat-number">{total_test_executions}</div>
                        <div class="stat-label">Total Test Executions</div>
                        <div class="stat-sublabel">Across All Test Cases</div>
                    </div>
                    <div class="stat-card failed">
                        <div class="stat-icon"><i class="fas fa-exclamation-triangle"></i></div>
                        <div class="stat-number">{tests_with_failures}</div>
                        <div class="stat-label">Tests with Failures</div>
                        <div class="stat-sublabel">Need Attention</div>
                    </div>
                    <div class="stat-card not-app">
                        <div class="stat-icon"><i class="fas fa-info-circle"></i></div>
                        <div class="stat-number">{tests_with_not_app}</div>
                        <div class="stat-label">Tests with N/A</div>
                        <div class="stat-sublabel">Non-UBI Based</div>
                    </div>
                </div>
                
                <div class="table-container">
                    <table class="test-analysis-table">
                        <thead>
                            <tr>
                                <th><i class="fas fa-clipboard-list"></i> Test Case</th>
                                <th><i class="fas fa-hashtag"></i> Total Tests</th>
                                <th><i class="fas fa-check-circle"></i> Passed</th>
                                <th><i class="fas fa-times-circle"></i> Failed</th>
                                <th><i class="fas fa-info-circle"></i> Not Applicable</th>
                                <th><i class="fas fa-chart-line"></i> Success Rate</th>
                            </tr>
                        </thead>
                        <tbody>"""

    # Define test case icons
    test_case_icons = {
        'BasedOnUbi': 'fas fa-cube',
        'HasLicense': 'fas fa-file-contract',
        'HasModifiedFiles': 'fas fa-file-edit',
        'HasNoProhibitedLabels': 'fas fa-tags',
        'HasNoProhibitedPackages': 'fas fa-box',
        'HasProhibitedContainerName': 'fas fa-signature',
        'HasRequiredLabel': 'fas fa-tag',
        'HasUniqueTag': 'fas fa-fingerprint',
        'LayerCountAcceptable': 'fas fa-layer-group',
        'RunAsNonRoot': 'fas fa-user-shield'
    }

    # Add test case analysis rows with enhanced styling
    for test_case, stats in sorted(test_case_stats.items()):
        success_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
        
        # Get icon for test case
        icon = test_case_icons.get(test_case, 'fas fa-check')
        
        # Determine success rate category and color
        if success_rate >= 95:
            rate_class = 'excellent'
        elif success_rate >= 80:
            rate_class = 'good'
        elif success_rate >= 60:
            rate_class = 'warning'
        else:
            rate_class = 'poor'
        
        # Format numbers with color coding
        total_display = f'<span class="metric-number metric-total">{stats["total"]}</span>'
        
        passed_display = f'<span class="metric-number metric-passed">{stats["passed"]}</span>' if stats['passed'] > 0 else f'<span class="metric-number metric-zero">0</span>'
        
        failed_display = f'<span class="metric-number metric-failed">{stats["failed"]}</span>' if stats['failed'] > 0 else f'<span class="metric-number metric-zero">0</span>'
        
        not_app_display = f'<span class="metric-number metric-not-app">{stats["not_app"]}</span>' if stats['not_app'] > 0 else f'<span class="metric-number metric-zero">0</span>'
        
        # Create success rate with progress bar
        success_rate_display = f'''
            <div class="success-rate-container">
                <div class="success-rate-bar">
                    <div class="success-rate-fill {rate_class}" style="width: {success_rate}%"></div>
                </div>
                <div class="success-rate-text {rate_class}">{success_rate:.1f}%</div>
            </div>
        '''
        
        html_content += f"""
                            <tr>
                                <td>
                                    <div class="test-case-name">
                                        <i class="{icon} test-case-icon"></i>
                                        {test_case}
                                    </div>
                                </td>
                                <td>{total_display}</td>
                                <td>{passed_display}</td>
                                <td>{failed_display}</td>
                                <td>{not_app_display}</td>
                                <td>{success_rate_display}</td>
                            </tr>"""

    html_content += """
                        </tbody>
                    </table>
                </div>
            </div>"""

    # Add Failed Tests section if there are failures
    if failed_images:
        html_content += """
            <!-- Failed Tests Details -->
            <div class="section" id="failed-tests-details">
                <h2>⚠️ Failed Tests Details</h2>
                <div class="image-list">
                    <h3>Images with Failed Tests:</h3>"""

        # Show failed images and their failed tests
        for image_name, failed_tests in failed_images.items():
            html_content += f"""
                    <div class="image-item">
                        <h4>{image_name}</h4>
                        <p><strong>Failed Tests ({len(failed_tests)}):</strong></p>
                        <ul>"""
            for test in failed_tests:
                html_content += f"""
                            <li><span class="status-badge status-failed">{test}</span> - Test failed for this image</li>"""
            html_content += """
                        </ul>
                    </div>"""

        # Show grouped failure analysis
        if len(failed_test_details) > 1:
            html_content += """
                    <h3>Failure Analysis by Test Type:</h3>"""
            for test_case, affected_images in failed_test_details.items():
                if len(affected_images) > 1:
                    html_content += f"""
                    <div class="image-item">
                        <h4>{test_case} Failures ({len(affected_images)} images)</h4>
                        <p>The following images failed the {test_case} test:</p>
                        <ul>"""
                    for img in affected_images[:10]:  # Limit to first 10
                        html_content += f"""
                            <li>{img}</li>"""
                    if len(affected_images) > 10:
                        html_content += f"""
                            <li><em>... and {len(affected_images) - 10} more images</em></li>"""
                    html_content += """
                        </ul>
                    </div>"""

        html_content += """
                </div>
            </div>"""

    # Add Image Categories section - adaptive approach
    if adaptive_categories and len(adaptive_categories) > 1:
        html_content += """
            <!-- Image Categories -->
            <div class="section">
                <h2>🏗️ Image Categories</h2>
                <div class="stats-grid">"""
        
        # Add dynamic category cards (limit to first 4 for layout)
        for i, (category_name, images) in enumerate(list(adaptive_categories.items())[:4]):
            html_content += f"""
                    <div class="stat-card">
                        <div class="stat-number">{len(images)}</div>
                        <div class="stat-label">{category_name}</div>
                    </div>"""
        
        html_content += """
                </div>"""
        
        # Show additional categories if there are more than 4
        if len(adaptive_categories) > 4:
            html_content += """
                <div class="stats-grid">"""
            for category_name, images in list(adaptive_categories.items())[4:8]:  # Show next 4
                html_content += f"""
                    <div class="stat-card">
                        <div class="stat-number">{len(images)}</div>
                        <div class="stat-label">{category_name}</div>
                    </div>"""
            html_content += """
                </div>"""
        
        # Always show the health summary
        html_content += f"""
                <div class="stats-grid">
                    <div class="stat-card passed">
                        <div class="stat-number">{images_perfect}</div>
                        <div class="stat-label">Images with Perfect Score</div>
                    </div>
                    <div class="stat-card failed">
                        <div class="stat-number">{images_with_issues}</div>
                        <div class="stat-label">Images with Issues</div>
                    </div>
                </div>
            </div>"""
    else:
        # Fallback: Only show health summary if no meaningful categories found
        html_content += f"""
            <!-- Image Summary -->
            <div class="section">
                <h2>📊 Image Summary</h2>
                <div class="stats-grid">
                    <div class="stat-card">
                        <div class="stat-number">{total_unique_images}</div>
                        <div class="stat-label">Total Unique Images</div>
                    </div>
                    <div class="stat-card passed">
                        <div class="stat-number">{images_perfect}</div>
                        <div class="stat-label">Images with Perfect Score</div>
                    </div>
                    <div class="stat-card failed">
                        <div class="stat-number">{images_with_issues}</div>
                        <div class="stat-label">Images with Issues</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{images_with_issues/total_unique_images*100 if total_unique_images > 0 else 0:.1f}%</div>
                        <div class="stat-label">Issues Rate</div>
                    </div>
                </div>
            </div>"""

    # Add Recommendations section
    recommendations = []
    if failed_count > 0:
        recommendations.append("Address all failed security tests to improve overall security posture")
        
    if 'RunAsNonRoot' in failed_test_details:
        recommendations.append(f"Configure {len(failed_test_details['RunAsNonRoot'])} images to run as non-root users for better security")
        
    if 'BasedOnUbi' in failed_test_details:
        recommendations.append("Standardize base images to use approved UBI (Universal Base Image)")
        
    if 'HasLicense' in failed_test_details:
        recommendations.append("Ensure all images have proper license information for compliance")

    if recommendations or failed_count == 0:
        html_content += """
            <!-- Recommendations -->
            <div class="section">
                <h2>💡 Recommendations</h2>
                <div class="summary-box">"""
        
        if failed_count == 0:
            html_content += """
                    <h3>✅ Excellent Results!</h3>
                    <ul>
                        <li>All security tests passed successfully</li>
                        <li>Continue maintaining current security standards</li>
                        <li>Regular security scans should be performed to maintain compliance</li>
                        <li>Consider implementing automated scanning in CI/CD pipeline</li>
                    </ul>"""
        else:
            html_content += """
                    <h3>Priority Actions:</h3>
                    <ul>"""
            for recommendation in recommendations:
                html_content += f"""
                        <li>{recommendation}</li>"""
            html_content += """
                    </ul>
                    
                    <h3>Security Best Practices:</h3>
                    <ul>
                        <li>Implement non-root user configuration for all containers</li>
                        <li>Add required security labels to all images</li>
                        <li>Ensure consistent licensing across all components</li>
                        <li>Regular security scanning and vulnerability assessments</li>
                        <li>Keep base images updated with latest security patches</li>
                    </ul>"""
        
        html_content += """
                </div>
            </div>"""

    # Add detailed results table
    html_content += """
            <!-- Detailed Results -->
            <div class="section" id="detailed-test-results">
                <h2><i class="section-icon fas fa-list-alt"></i>Detailed Test Results</h2>
                <div class="table-container">
                    <table class="searchable-table" id="resultsTable">
                        <thead>
                            <tr>
                                <th>Image Name <i class="fas fa-sort"></i></th>
                                <th>Image Tag <i class="fas fa-sort"></i></th>
                                <th>Test Case <i class="fas fa-sort"></i></th>
                                <th>Status <i class="fas fa-sort"></i></th>
                                <th>Modified Files</th>
                            </tr>
                        </thead>
                        <tbody>"""

    # Add all test results with enhanced tooltips for NOT_APP
    for row in dict_data:
        status_class = {
            'PASSED': 'status-passed',
            'FAILED': 'status-failed',
            'NOT_APP': 'status-not-app'
        }.get(row['Status'], 'status-not-app')
        
        # Create tooltip for NOT_APP status
        tooltip_text = ""
        status_icon = ""
        if row['Status'] == 'NOT_APP':
            tooltip_text = ' data-tooltip="This test is not applicable because the image is not based on Universal Base Image (UBI). UBI-based images are required for certain security validations."'
            status_icon = '<i class="fas fa-question-circle"></i> '
        elif row['Status'] == 'FAILED':
            status_icon = '<i class="fas fa-times-circle"></i> '
        elif row['Status'] == 'PASSED':
            status_icon = '<i class="fas fa-check-circle"></i> '
        
        modified_files = row.get('Has Modified Files', '').replace(':', '<br>') if row.get('Has Modified Files') else ''
        
        html_content += f"""
                            <tr data-image="{row['Image Name'].lower()}" data-test="{row['Test Case'].lower()}" data-status="{row['Status'].lower()}">
                                <td>
                                    <strong>{row['Image Name']}</strong>
                                    <button class="copy-btn" onclick="copyToClipboard('{row['Image Name']}')" title="Copy image name">
                                        <i class="fas fa-copy"></i>
                                    </button>
                                </td>
                                <td><code>{row['Image Tag']}</code></td>
                                <td>{row['Test Case']}</td>
                                <td>
                                    <span class="status-badge {status_class} tooltip"{tooltip_text}>
                                        {status_icon}{row['Status']}
                                    </span>
                                </td>
                                <td>{modified_files if modified_files else '<em>None</em>'}</td>
                            </tr>"""

    html_content += f"""
                        </tbody>
                    </table>
                </div>
            </div>
            
            <div class="timestamp">
                <i class="fas fa-calendar-alt"></i> Report generated on: {current_time}
            </div>
        </div>
    </div>

    <script>
        // Initialize page
        document.addEventListener('DOMContentLoaded', function() {{
            // Hide loading spinner
            document.getElementById('loadingSpinner').style.display = 'none';
            
            // Initialize search functionality
            initializeSearch();
            
            // Initialize table sorting
            initializeTableSorting();
            
            // Add smooth scrolling
            document.querySelectorAll('a[href^="#"]').forEach(anchor => {{
                anchor.addEventListener('click', function (e) {{
                    e.preventDefault();
                    document.querySelector(this.getAttribute('href')).scrollIntoView({{
                        behavior: 'smooth'
                    }});
                }});
            }});
        }});

        // Search functionality
        function initializeSearch() {{
            const searchInput = document.getElementById('searchInput');
            const table = document.getElementById('resultsTable');
            const rows = table.querySelectorAll('tbody tr');
            
            searchInput.addEventListener('input', function() {{
                const searchTerm = this.value.toLowerCase().trim();
                let visibleRows = 0;
                
                rows.forEach(row => {{
                    const imageData = row.getAttribute('data-image') || '';
                    const testData = row.getAttribute('data-test') || '';
                    const statusData = row.getAttribute('data-status') || '';
                    const textContent = row.textContent.toLowerCase();
                    
                    if (searchTerm === '' || 
                        imageData.includes(searchTerm) || 
                        testData.includes(searchTerm) || 
                        statusData.includes(searchTerm) ||
                        textContent.includes(searchTerm)) {{
                        row.style.display = '';
                        row.classList.remove('hidden');
                        visibleRows++;
                    }} else {{
                        row.style.display = 'none';
                        row.classList.add('hidden');
                    }}
                }});
                
                // Show search results count
                updateSearchResults(visibleRows, rows.length);
            }});
        }}

        // Filter table by status (for clickable stat cards)
        function filterTableByStatus(status) {{
            // First scroll to the table
            setTimeout(() => {{
                const table = document.getElementById('resultsTable');
                if (table) {{
                    // Clear any existing search
                    const searchInput = document.getElementById('searchInput');
                    searchInput.value = '';
                    
                    // Apply status filter
                    const rows = table.querySelectorAll('tbody tr');
                    let visibleRows = 0;
                    
                    rows.forEach(row => {{
                        const statusData = row.getAttribute('data-status') || '';
                        let shouldShow = false;
                        
                        if (status === 'passed' && statusData === 'passed') {{
                            shouldShow = true;
                        }} else if (status === 'failed' && statusData === 'failed') {{
                            shouldShow = true;
                        }} else if (status === 'not_app' && statusData === 'not_app') {{
                            shouldShow = true;
                        }}
                        
                        if (shouldShow) {{
                            row.style.display = '';
                            row.classList.remove('hidden');
                            visibleRows++;
                        }} else {{
                            row.style.display = 'none';
                            row.classList.add('hidden');
                        }}
                    }});
                    
                    // Update search results count
                    updateSearchResults(visibleRows, rows.length);
                    
                    // Show filter indicator
                    showFilterIndicator(status, visibleRows);
                }}
            }}, 300); // Small delay to allow smooth scrolling
        }}
        
        // Show filter indicator
        function showFilterIndicator(status, count) {{
            const statusLabels = {{
                'passed': 'Passed Tests',
                'failed': 'Failed Tests', 
                'not_app': 'Not Applicable Tests'
            }};
            
            showToast(`Filtered to show ${{count}} ${{statusLabels[status] || 'results'}}`, 'success');
            
            // Add clear filter button to search area
            let clearButton = document.querySelector('.clear-filter-btn');
            if (!clearButton) {{
                clearButton = document.createElement('button');
                clearButton.className = 'btn clear-filter-btn';
                clearButton.innerHTML = '<i class="fas fa-times"></i> Clear Filter';
                clearButton.onclick = clearTableFilter;
                clearButton.style.marginLeft = '8px';
                document.querySelector('.action-buttons').insertAdjacentElement('afterbegin', clearButton);
            }}
        }}
        
        // Clear table filter
        function clearTableFilter() {{
            const searchInput = document.getElementById('searchInput');
            searchInput.value = '';
            searchInput.dispatchEvent(new Event('input'));
            
            const clearButton = document.querySelector('.clear-filter-btn');
            if (clearButton) {{
                clearButton.remove();
            }}
            
            showToast('Filter cleared - showing all results');
        }}
        
        function updateSearchResults(visible, total) {{
            let statusElement = document.querySelector('.search-status');
            if (!statusElement) {{
                statusElement = document.createElement('div');
                statusElement.className = 'search-status';
                statusElement.style.cssText = 'margin-top: 8px; font-size: 0.875rem; color: var(--gray-600);';
                document.querySelector('.search-box').appendChild(statusElement);
            }}
            
            if (visible === total) {{
                statusElement.textContent = '';
            }} else {{
                statusElement.textContent = `Showing ${{visible}} of ${{total}} results`;
            }}
        }}

        // Table sorting functionality
        function initializeTableSorting() {{
            const headers = document.querySelectorAll('#resultsTable th');
            headers.forEach((header, index) => {{
                if (header.querySelector('.fa-sort')) {{
                    header.style.cursor = 'pointer';
                    header.addEventListener('click', () => sortTable(index));
                }}
            }});
        }}
        
        function sortTable(columnIndex) {{
            const table = document.getElementById('resultsTable');
            const tbody = table.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr'));
            const header = table.querySelectorAll('th')[columnIndex];
            const icon = header.querySelector('i');
            
            // Reset all other sorting icons
            table.querySelectorAll('th i').forEach(i => {{
                if (i !== icon) i.className = 'fas fa-sort';
            }});
            
            // Determine sort direction
            const isAsc = icon.classList.contains('fa-sort-up');
            const newDirection = isAsc ? 'down' : 'up';
            icon.className = `fas fa-sort-${{newDirection}}`;
            
            // Sort rows
            rows.sort((a, b) => {{
                const aText = a.children[columnIndex].textContent.trim();
                const bText = b.children[columnIndex].textContent.trim();
                
                // Handle numeric sorting for status
                if (columnIndex === 3) {{ // Status column
                    const statusOrder = {{'FAILED': 0, 'NOT_APP': 1, 'PASSED': 2}};
                    const aVal = statusOrder[aText] || 999;
                    const bVal = statusOrder[bText] || 999;
                    return newDirection === 'up' ? aVal - bVal : bVal - aVal;
                }}
                
                // Text sorting
                const result = aText.localeCompare(bText);
                return newDirection === 'up' ? result : -result;
            }});
            
            // Reorder DOM elements
            rows.forEach(row => tbody.appendChild(row));
        }}

        // Copy to clipboard functionality
        function copyToClipboard(text) {{
            navigator.clipboard.writeText(text).then(function() {{
                // Show success feedback
                const event = window.event;
                if (event && event.target) {{
                    const btn = event.target.closest('.copy-btn');
                    if (btn) {{
                        const originalClass = btn.className;
                        btn.className = 'copy-btn copied';
                        btn.innerHTML = '<i class="fas fa-check"></i>';
                        setTimeout(() => {{
                            btn.className = originalClass;
                            btn.innerHTML = '<i class="fas fa-copy"></i>';
                        }}, 2000);
                    }}
                }}
                
                // Show toast notification
                showToast(`Copied: ${{text}}`);
            }}).catch(function(err) {{
                console.error('Failed to copy: ', err);
                showToast('Failed to copy to clipboard', 'error');
            }});
        }}

        // Toast notifications
        function showToast(message, type = 'success') {{
            const toast = document.createElement('div');
            toast.className = `toast toast-${{type}}`;
            toast.style.cssText = `
                position: fixed;
                top: 20px;
                right: 20px;
                background: ${{type === 'success' ? 'var(--success-color)' : 'var(--error-color)'}};
                color: white;
                padding: 12px 20px;
                border-radius: 8px;
                box-shadow: var(--shadow-lg);
                z-index: 10000;
                animation: slideInRight 0.3s ease;
            `;
            toast.textContent = message;
            
            document.body.appendChild(toast);
            setTimeout(() => {{
                toast.style.animation = 'slideOutRight 0.3s ease';
                setTimeout(() => toast.remove(), 300);
            }}, 3000);
        }}

        // Export functionality
        function exportToCSV() {{
            const table = document.getElementById('resultsTable');
            const rows = table.querySelectorAll('tr');
            let csvContent = '';
            
            rows.forEach(row => {{
                const cells = row.querySelectorAll('th, td');
                const rowData = Array.from(cells).map(cell => {{
                    // Clean up cell content
                    let content = cell.textContent.trim();
                    content = content.replace(/\\s+/g, ' '); // Normalize whitespace
                    return `"${{content.replace(/"/g, '""')}}"`; // Escape quotes
                }});
                csvContent += rowData.join(',') + '\\n';
            }});
            
            // Download CSV
            const blob = new Blob([csvContent], {{ type: 'text/csv;charset=utf-8;' }});
            const link = document.createElement('a');
            link.href = URL.createObjectURL(blob);
            link.download = `container-security-report-${{new Date().toISOString().split('T')[0]}}.csv`;
            link.click();
            
            showToast('CSV export completed successfully!');
        }}

        // Print functionality
        function printReport() {{
            window.print();
        }}

        // Share functionality
        function copyReportLink() {{
            const url = window.location.href;
            copyToClipboard(url);
            showToast('Report link copied to clipboard!');
        }}

        // Keyboard shortcuts
        document.addEventListener('keydown', function(e) {{
            // Ctrl/Cmd + F for search
            if ((e.ctrlKey || e.metaKey) && e.key === 'f') {{
                e.preventDefault();
                document.getElementById('searchInput').focus();
            }}
            
            // Ctrl/Cmd + P for print
            if ((e.ctrlKey || e.metaKey) && e.key === 'p') {{
                e.preventDefault();
                printReport();
            }}
            
            // Escape to clear search
            if (e.key === 'Escape') {{
                const searchInput = document.getElementById('searchInput');
                if (searchInput.value) {{
                    searchInput.value = '';
                    searchInput.dispatchEvent(new Event('input'));
                }}
            }}
        }});

        // Add CSS animations
        const style = document.createElement('style');
        style.textContent = `
            @keyframes slideInRight {{
                from {{ transform: translateX(100%); opacity: 0; }}
                to {{ transform: translateX(0); opacity: 1; }}
            }}
            @keyframes slideOutRight {{
                from {{ transform: translateX(0); opacity: 1; }}
                to {{ transform: translateX(100%); opacity: 0; }}
            }}
        `;
        document.head.appendChild(style);
    </script>
</body>
</html>"""

    # Write HTML file
    try:
        with open(output_html, 'w', encoding='utf-8') as f:
            f.write(html_content)
    except Exception as e:
        raise Exception(f"Failed to write HTML report: {e}")

# ------------------------------------------------------------------------------
# XLSX writing function
# ------------------------------------------------------------------------------

def write_and_format_xlsx(data: List[List[str]], detailed_checks: List[Dict[str, Any]], output_xlsx: str) -> None:
    """
    Takes scan result data and detailed check information, sorts it by 'Status' and 'Test Case' (with custom order),
    formats the worksheet, and saves the result as an Excel workbook with two sheets.
    """
    if not data:
        raise ValueError("No data to write to XLSX")
    
    # Convert to dictionary format for easier processing
    headers = ["Image Name", "Image Tag", "Has Modified Files", "Test Case", "Status"]
    dict_data = []
    for row in data:
        if len(row) >= 5:
            dict_data.append({
                headers[0]: row[0],
                headers[1]: row[1], 
                headers[2]: row[2],
                headers[3]: row[3],
                headers[4]: row[4]
            })
    
    # Sort data by Status and Test Case with custom order
    status_order = {'FAILED': 0, 'NOT_APP': 1, 'PASSED': 2}
    dict_data.sort(key=lambda x: (status_order.get(x.get('Status', ''), 3), x.get('Test Case', '')))
    
    # Create workbook and worksheets
    wb = Workbook()
    
    # First sheet - Summary
    ws_summary = wb.active
    if ws_summary is None:
        raise ValueError("Failed to create worksheet")
    ws_summary.title = "Summary"
    
    # Set column widths for summary sheet
    column_widths = {
        'A': 20, 'B': 30, 'C': 40, 'D': 30, 'E': 20
    }
    for col, width in column_widths.items():
        ws_summary.column_dimensions[col].width = width
    
    # Write headers to summary sheet
    ws_summary.append(headers)
    
    # Write data rows to summary sheet
    for row_data in dict_data:
        ws_summary.append([row_data.get(header, '') for header in headers])
    
    # Enable text wrap for column C (Has Modified Files)
    for cell in ws_summary['C']:
        cell.alignment = Alignment(wrap_text=True)
    
    # Format the Status column (assumed to be column E)
    status_colors = {
        'PASSED': '006400',    # Dark green
        'FAILED': 'FF0000',    # Red
        'NOT_APP': 'FFA500'    # Dark orange
    }
    
    for row in ws_summary.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell_value = str(cell.value) if cell.value is not None else ""
            if cell_value in status_colors:
                cell.font = Font(color=status_colors[cell_value])
    
    # Set alignment: center for "Status" and "Image Tag", left for others
    for col in ws_summary.columns:
        col_list = list(col)
        if col_list:
            header_value = str(col_list[0].value) if col_list[0].value is not None else ""
            if header_value in ['Status', 'Image Tag']:
                for cell in col_list:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                for cell in col_list:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Format header row for summary sheet
    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Second sheet - Detailed Checks
    ws_details = wb.create_sheet(title="Detailed Checks")
    
    # Headers for detailed checks sheet
    detail_headers = ["Image Name", "Image Tag", "Check Name", "Elapsed Time", "Description", "Help", "Suggestion", "Knowledge Base URL", "Check URL"]
    
    # Set column widths for detailed checks sheet
    detail_column_widths = {
        'A': 20,  # Image Name
        'B': 15,  # Image Tag
        'C': 25,  # Check Name
        'D': 12,  # Elapsed Time
        'E': 50,  # Description
        'F': 50,  # Help
        'G': 50,  # Suggestion
        'H': 40,  # Knowledge Base URL
        'I': 40   # Check URL
    }
    for col_letter, width in detail_column_widths.items():
        ws_details.column_dimensions[col_letter].width = width
    
    # Write headers to detailed checks sheet
    ws_details.append(detail_headers)
    
    # Write detailed check data
    if detailed_checks:
        for check in detailed_checks:
            row_data = [
                check.get('image_name', ''),
                check.get('image_tag', ''),
                check.get('name', ''),
                check.get('elapsed_time', ''),
                check.get('description', ''),
                check.get('help', ''),
                check.get('suggestion', ''),
                check.get('knowledgebase_url', ''),
                check.get('check_url', '')
            ]
            ws_details.append(row_data)
    
    # Enable text wrap for description, help, and suggestion columns
    for col in ['E', 'F', 'G']:
        for cell in ws_details[col]:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Set alignment for detailed checks sheet
    for col in ws_details.columns:
        col_list = list(col)
        if col_list:
            for cell in col_list:
                cell.alignment = Alignment(horizontal='left', vertical='top')
    
    # Format header row for detailed checks sheet
    for cell in ws_details[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    wb.save(output_xlsx)

# ------------------------------------------------------------------------------
# Legacy CSV-to-XLSX conversion function (kept for backward compatibility)
# ------------------------------------------------------------------------------

def convert_and_format_csv_to_xlsx(input_csv: str, output_xlsx: str) -> None:
    """
    Reads the CSV file, sorts it by 'Status' and 'Test Case' (with custom order),
    formats the worksheet, and saves the result as an Excel workbook.
    Note: This legacy function creates only the summary sheet.
    """
    # Read CSV data manually instead of using pandas
    with open(input_csv, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        data = list(reader)
    
    # Sort data by Status and Test Case with custom order
    status_order = {'FAILED': 0, 'NOT_APP': 1, 'PASSED': 2}
    data.sort(key=lambda x: (status_order.get(x.get('Status', ''), 3), x.get('Test Case', '')))
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    
    if ws is None:
        raise ValueError("Failed to create worksheet")
    
    # Set column widths
    column_widths = {
        'A': 20, 'B': 30, 'C': 40, 'D': 30, 'E': 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Write headers
    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        
        # Write data rows
        for row in data:
            ws.append([row.get(header, '') for header in headers])
    
    # Enable text wrap for column C (Has Modified Files)
    for cell in ws['C']:
        cell.alignment = Alignment(wrap_text=True)
    
    # Format the Status column (assumed to be column E)
    status_colors = {
        'PASSED': '006400',    # Dark green
        'FAILED': 'FF0000',    # Red
        'NOT_APP': 'FFA500'    # Dark orange
    }
    
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell_value = str(cell.value) if cell.value is not None else ""
            if cell_value in status_colors:
                cell.font = Font(color=status_colors[cell_value])
    
    # Set alignment: center for "Status" and "Image Tag", left for others
    for col in ws.columns:
        col_list = list(col)
        if col_list:
            header_value = str(col_list[0].value) if col_list[0].value is not None else ""
            if header_value in ['Status', 'Image Tag']:
                for cell in col_list:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                for cell in col_list:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Format header row
    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    wb.save(output_xlsx)

# ------------------------------------------------------------------------------
# PreflightScanner class
# ------------------------------------------------------------------------------

class PreflightScanner:
    """Main class for handling container image scanning with preflight."""
    
    RESULT_XLSX = "images_scan_results.xlsx"
    RESULT_HTML = "image-scanning-report.html"
    MIN_PYTHON_VERSION = (3, 8)
    MIN_PREFLIGHT_VERSION = "1.6.11"

    def __init__(self):
        self.args = self.parse_args()
        self.api_token = self.args.api_token
        self.repo_namespace = self.args.repo_namespace
        self.cnf_prefix = self.args.cnf_prefix
        self.tag_type = self.args.tag_type or "name"
        self.auth_json = self.args.auth_json
        self.fqdn = self.args.fqdn
        self.filter = self.args.filter or "chartrepo"
        self.image_file = self.args.image_file
        self.parallel = self.args.parallel
        self.debug = self.args.debug
        self.image_list: List[str] = []

    def parse_args(self) -> argparse.Namespace:
        """Parse command line arguments."""
        parser = argparse.ArgumentParser(
            description="""Scan container images using preflight in parallel and write results to both XLSX and HTML reports.

Usage Examples:

API-based:
  ./quick_scan_container_images_parallel.py --repo-namespace avareg_5gc --cnf-prefix "global-|specific" \\
      --auth-json auth.json --api-token xxxxxx --fqdn quay.io --tag-type name --filter "existed_image|tested_image" --parallel 2

Offline:
  ./quick_scan_container_images_parallel.py --image-file image_list.txt --auth-json auth.json --parallel 2
  ./quick_scan_container_images_parallel.py --image-file image_list.txt
  ./quick_scan_container_images_parallel.py --image-file image_list.txt --parallel 2

Note: if preflight scan failed for some reason, then you add --debug
""",
            formatter_class=argparse.RawTextHelpFormatter
        )
        
        # API-based arguments
        parser.add_argument("--repo-namespace", "-rn", 
                          help="Repository namespace (e.g., 'avareg_5gc' or 'avu0').")
        parser.add_argument("--cnf-prefix", "-cp", 
                          help="CNF image prefix to search for (e.g., 'global-' or 'global|non-global').")
        parser.add_argument("--tag-type", "-t", 
                          help="Image tag type: 'name' (default) or 'digest'.")
        parser.add_argument("--api-token", "-at", 
                          help="API token (Bearer Token) for registry access.")
        parser.add_argument("--auth-json", "-d", 
                          help="Path to Docker authentication JSON file (if required).")
        
        # Offline argument
        parser.add_argument("--image-file", "-img", 
                          help="Text file with a list of images (one per line).")
        
        # Common arguments
        parser.add_argument("--fqdn", "-fq", 
                          help="Fully-qualified domain name of your registry (e.g., 'quay.io').")
        parser.add_argument("--filter", "-ft", 
                          help="Filter to exclude images (e.g., 'existed_image|tested_image').")
        parser.add_argument("--parallel", "-p", type=int, default=1, 
                          help="Number of images to scan in parallel (default: 1).")
        parser.add_argument("--debug", action="store_true", 
                          help="Enable debug logging.")
        
        return parser.parse_args()

    def log(self, message: str) -> None:
        """Log a message with timestamp."""
        timestamp = datetime.datetime.now().strftime("%Y%m%d-%H:%M:%S")
        print(f"{timestamp} {message}")

    def exit_with_error(self, message: str) -> None:
        """Log error message and exit."""
        self.log(message)
        sys.exit(1)

    @staticmethod
    def file_exists(filepath: str) -> bool:
        """Check if a file exists."""
        return os.path.exists(filepath)

    def rename_file(self, old_name: str, new_name: str) -> bool:
        """Rename a file safely."""
        if not self.file_exists(old_name):
            self.log(f"Error: file '{old_name}' does not exist")
            return False
        try:
            os.rename(old_name, new_name)
            self.log(f"File '{old_name}' has been renamed to '{new_name}'")
            return True
        except Exception as e:
            self.log(f"Error renaming file: {e}")
            return False

    def print_status(self, description: str, status: str, color: str = "32") -> None:
        """Print formatted status message."""
        if status == "OK":
            color_code = "32"  # Green
        elif status == "NOK":
            color_code = "31"  # Red
        else:
            color_code = "33"  # Yellow
        
        print(f"{description:<48} \033[1;{color_code}m{status:<24}\033[m")

    def check_required_tools(self) -> None:
        """Check if required tools are installed."""
        if shutil.which("python3") and shutil.which("preflight"):
            self.print_status("python3 and preflight installed", "OK")
        else:
            self.print_status("python3 and preflight installed", "NOK")
            sys.exit(1)

    def check_python_version(self) -> None:
        """Check if Python version meets minimum requirements."""
        current_version = f"{sys.version_info.major}.{sys.version_info.minor}"
        min_version = f"{self.MIN_PYTHON_VERSION[0]}.{self.MIN_PYTHON_VERSION[1]}"
        
        if sys.version_info >= self.MIN_PYTHON_VERSION:
            self.print_status(f"Python3 version ({current_version}>={min_version})", "OK")
        else:
            self.print_status(f"Python3 version ({current_version}>={min_version})", "NOK")
            sys.exit(1)

    @staticmethod
    def version_tuple(version_str: str) -> tuple:
        """Convert version string to tuple for comparison."""
        return tuple(map(int, version_str.split(".")))

    def check_preflight_version(self) -> None:
        """Check if preflight version meets minimum requirements."""
        try:
            result = subprocess.run(
                ["preflight", "--version"], 
                capture_output=True, text=True, check=True
            )
            match = re.search(r'(\d+\.\d+\.\d+)', result.stdout)
            if match:
                current_version = match.group(1)
                if self.version_tuple(current_version) < self.version_tuple(self.MIN_PREFLIGHT_VERSION):
                    self.print_status(f"Preflight version ({current_version}>={self.MIN_PREFLIGHT_VERSION})", "NOK")
                    sys.exit(1)
                else:
                    self.print_status(f"Preflight version ({current_version}>={self.MIN_PREFLIGHT_VERSION})", "OK")
            else:
                self.log("Could not determine preflight version.")
                sys.exit(1)
        except subprocess.CalledProcessError as e:
            self.log(f"Error checking preflight version: {e}")
            sys.exit(1)

    def check_python_dependencies(self) -> None:
        """Check if required Python dependencies are installed."""
        try:
            import openpyxl  # noqa: F401
            self.print_status("Python Openpyxl installed", "OK")
        except ImportError as e:
            missing = str(e).split("No module named")[-1].strip(" '")
            self.print_status(f"Python {missing}", "NOK")
            sys.exit(1)

    def check_registry_connection(self) -> None:
        """Check connection to the registry."""
        if not self.fqdn:
            return
            
        if shutil.which("nc"):
            try:
                subprocess.run(
                    ["nc", "-zv4", self.fqdn, "80"], 
                    capture_output=True, text=True, check=True
                )
                self.print_status(f"{self.fqdn} connection", "OK")
            except subprocess.CalledProcessError:
                self.print_status(f"{self.fqdn} connection", "NOK")
                sys.exit(1)
        else:
            self.print_status(f"{self.fqdn} connection", "SKIPPED")

    def check_registry_authentication(self) -> None:
        """Check registry authentication using Bearer token."""
        if not self.api_token or not self.fqdn or not self.repo_namespace:
            return
            
        url = f"https://{self.fqdn}/api/v1/repository?namespace={self.repo_namespace}"
        try:
            result = subprocess.run([
                "curl", "-I", "--silent", "-o", "/dev/null", "-w", "%{http_code}",
                "-X", "GET", "-H", f"Authorization: Bearer {self.api_token}", url
            ], capture_output=True, text=True, check=True)
            
            if result.stdout.strip() == "200":
                self.print_status("Registry auth (Bearer Token)", "OK")
            else:
                self.print_status("Registry auth (Bearer Token)", "NOK")
                sys.exit(1)
        except subprocess.CalledProcessError as e:
            self.log(f"Error checking registry authentication: {e}")
            sys.exit(1)

    def ensure_trailing_newline(self, filepath: str) -> None:
        """Ensure file ends with a newline."""
        try:
            with open(filepath, "rb+") as f:
                f.seek(-1, os.SEEK_END)
                if f.read(1) != b"\n":
                    f.write(b"\n")
        except Exception as e:
            self.log(f"Error ensuring trailing newline in {filepath}: {e}")

    def fetch_image_list_from_api(self) -> List[str]:
        """Fetch image list from registry API."""
        if not self.fqdn or not self.repo_namespace:
            return []
            
        try:
            api_url = f"https://{self.fqdn}/api/v1/repository?namespace={self.repo_namespace}"
            proc = subprocess.run([
                "curl", "--silent", "-X", "GET", 
                "-H", f"Authorization: Bearer {self.api_token}", api_url
            ], capture_output=True, text=True, check=True)
            
            data = json.loads(proc.stdout)
            filtered_list = []
            
            for repo in data.get("repositories", []):
                name = repo.get("name", "")
                if self.cnf_prefix and self.cnf_prefix in name and self.filter not in name:
                    filtered_list.append(name)
            
            return filtered_list
        except Exception as e:
            self.log(f"Error fetching repository list: {e}")
            sys.exit(1)

    def build_image_list(self) -> None:
        """Build the list of images to scan."""
        default_images = ['ava-core/univ-nf-ava', 'ava-core/univ-nf-alex']
        
        if not self.api_token:
            # Offline mode - read from file
            if not self.image_file:
                self.exit_with_error("Image file must be provided in offline mode")
            
            self.ensure_trailing_newline(self.image_file)
            with open(self.image_file, "r") as f:
                self.image_list = [line.strip() for line in f if line.strip()]
        else:
            # API mode
            if not self.cnf_prefix:
                self.image_list = default_images
            else:
                filtered_list = self.fetch_image_list_from_api()
                self.image_list = filtered_list + default_images
        
        if not self.image_list:
            self.exit_with_error("No images found. Check the API response or the image list file!")

    def write_results_to_reports(self, scan_data: List[List[str]], detailed_checks: List[Dict[str, Any]]) -> None:
        """Write scan results to both XLSX and HTML reports."""
        try:
            # Write XLSX report
            write_and_format_xlsx(scan_data, detailed_checks, self.RESULT_XLSX)
            self.log(f"XLSX report written to {self.RESULT_XLSX} successfully!")
            
            # Write HTML report
            write_html_report(scan_data, detailed_checks, self.RESULT_HTML)
            self.log(f"HTML report written to {self.RESULT_HTML} successfully!")
            
        except Exception as e:
            self.exit_with_error(f"Failed to write reports: {e}")

    def get_image_details(self, image: str) -> Dict[str, Any]:
        """Get image details either from file or API."""
        try:
            if not self.api_token:
                # Offline mode
                image_details = image.strip()
                parts = image_details.split("/", 1)
                repo_img_tag = parts[1] if len(parts) > 1 else image_details
                img_name = repo_img_tag.split("/")[-1].split(":")[0]
                inspect_url = image_details
                # Extract tag from the rightmost colon that's not a port number
                # Tag is after the last colon, but only if it's not followed by a slash
                if ":" in image_details:
                    # Split by colon and check if the last part contains no slash (indicating it's a tag)
                    colon_parts = image_details.split(":")
                    if len(colon_parts) > 1 and "/" not in colon_parts[-1]:
                        tag = colon_parts[-1]
                    else:
                        tag = ""
                else:
                    tag = ""
            else:
                # API mode
                image_url = f"https://{self.fqdn}/api/v1/repository/{self.repo_namespace}/{image.strip()}"
                proc = subprocess.run([
                    "curl", "--silent", "-X", "GET", 
                    "-H", f"Authorization: Bearer {self.api_token}", image_url
                ], capture_output=True, text=True, check=True)
                
                data = json.loads(proc.stdout)
                
                if self.tag_type == "name":
                    tag_val = data["tags"][0]["name"] if data.get("tags") else ""
                    image_details = f"{data['name']}:{tag_val}"
                else:
                    tag_val = data["tags"][0]["manifest_digest"] if data.get("tags") else ""
                    image_details = f"{data['name']}@{tag_val}"
                
                tag = tag_val
                img_name = image.strip().split("/")[-1]
                repo_img_tag = f"{img_name}:{tag}"
                inspect_url = f"{self.fqdn}/{self.repo_namespace}/{image.strip()}:{tag}"
            
            return {
                "success": True,
                "image_details": image_details,
                "tag": tag,
                "img_name": img_name,
                "repo_img_tag": repo_img_tag,
                "inspect_url": inspect_url
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

    def parse_preflight_output(self, output: str, img_name: str, tag: str, temp_log_file: str) -> tuple:
        """Parse preflight output and extract results."""
        results = []
        detailed_checks = []
        
        for line in output.splitlines():
            m1 = re.search(r'check=([^ ]+)', line)
            m2 = re.search(r'result=([^ ]+)', line)
            if m1 and m2:
                results.append(f"{img_name},{m1.group(1)},{m2.group(1)}")
        
        if self.debug:
            self.log(f"Found {len(results)} basic check results for {img_name}")
        
        # Extract detailed check information from JSON output
        try:
            json_data = None
            
            # Try multiple approaches to find JSON data
            sources_to_check = [output]
            
            # Also check log file content
            try:
                with open(temp_log_file, "r") as lf:
                    log_content = lf.read()
                    sources_to_check.append(log_content)
            except Exception as e:
                if self.debug:
                    self.log(f"Could not read log file for JSON parsing: {e}")
            
            if self.debug:
                self.log(f"Python version: {sys.version_info}")
                self.log(f"Attempting JSON extraction for {img_name}")
                for i, source in enumerate(sources_to_check):
                    self.log(f"Source {i} length: {len(source)} characters")
            
            # Look for JSON in various formats
            for source in sources_to_check:
                if json_data:
                    break
                    
                # Try different JSON patterns (Python 3.8+ compatible)
                patterns = [
                    r'\{[^}]*"results"[^{]*\{[^}]*"passed"[^{]*\[[^]]*\{[^}]*"name"[^}]*\}[^]]*\][^}]*\}[^}]*\}',  # New format with results.passed
                    r'\{[^}]*"checks"[^{]*\[[^]]*\{[^}]*"name"[^}]*\}[^]]*\][^}]*\}',  # Old format with checks array
                    r'\{.*?"results".*?\}',  # Simple pattern for new format
                    r'\{.*?"checks".*?\}',  # Simple pattern for old format
                    r'(\{(?:[^{}]*|\{(?:[^{}]*|\{[^{}]*\})*\})*\})',  # Nested braces (Python 3.8+ compatible)
                ]
                
                for pattern in patterns:
                    matches = re.finditer(pattern, source, re.DOTALL)
                    for match in matches:
                        try:
                            potential_json = match.group(0)
                            parsed = json.loads(potential_json)
                            # Check for new format (results.passed/failed/errors) or old format (checks)
                            if ("results" in parsed and isinstance(parsed["results"], dict)) or \
                               ("checks" in parsed and isinstance(parsed["checks"], list)):
                                json_data = parsed
                                break
                        except json.JSONDecodeError:
                            continue
                    if json_data:
                        break
                
                # If no structured JSON found, try to find individual check objects
                if not json_data:
                    # Simple approach: look for lines that look like JSON objects
                    lines = source.split('\n')
                    temp_checks = []
                    
                    for line in lines:
                        line = line.strip()
                        if line.startswith('{') and '"name"' in line:
                            try:
                                # Try to parse the line as JSON
                                check_obj = json.loads(line)
                                if isinstance(check_obj, dict) and "name" in check_obj:
                                    temp_checks.append(check_obj)
                            except json.JSONDecodeError:
                                # Try to extract JSON-like structures with regex
                                match = re.search(r'\{[^}]*"name"[^}]*\}', line)
                                if match:
                                    try:
                                        check_obj = json.loads(match.group(0))
                                        if isinstance(check_obj, dict) and "name" in check_obj:
                                            temp_checks.append(check_obj)
                                    except json.JSONDecodeError:
                                        continue
                    
                    if temp_checks:
                        json_data = {"checks": temp_checks}
                        if self.debug:
                            self.log(f"Found {len(temp_checks)} checks using fallback method")
            
            # Extract check details from JSON - handle both old and new format
            if json_data:
                if self.debug:
                    self.log(f"Successfully parsed JSON data for {img_name}")
                checks_to_process = []
                
                # New format (preflight 1.14+): checks are under results.passed/failed/errors
                if "results" in json_data:
                    json_results = json_data["results"]  # Use different variable name to avoid collision
                    if self.debug:
                        self.log(f"Found new format JSON with results structure for {img_name}")
                    if "passed" in json_results:
                        checks_to_process.extend(json_results["passed"])
                        if self.debug:
                            self.log(f"Added {len(json_results['passed'])} passed checks")
                    if "failed" in json_results:
                        checks_to_process.extend(json_results["failed"])
                        if self.debug:
                            self.log(f"Added {len(json_results['failed'])} failed checks")
                    if "errors" in json_results:
                        checks_to_process.extend(json_results["errors"])
                        if self.debug:
                            self.log(f"Added {len(json_results['errors'])} error checks")
                
                # Old format: checks are directly under "checks"
                elif "checks" in json_data:
                    checks_to_process = json_data["checks"]
                
                # Process all found checks
                for check in checks_to_process:
                    if isinstance(check, dict):
                        detailed_check = {
                            'image_name': img_name,
                            'image_tag': tag,
                            'name': check.get('name', ''),
                            'elapsed_time': str(check.get('elapsed_time', '')),
                            'description': check.get('description', ''),
                            'help': check.get('help', ''),
                            'suggestion': check.get('suggestion', ''),
                            'knowledgebase_url': check.get('knowledgebase_url', ''),
                            'check_url': check.get('check_url', '')
                        }
                        detailed_checks.append(detailed_check)
            
            if self.debug:
                if detailed_checks:
                    self.log(f"Extracted {len(detailed_checks)} detailed checks for {img_name}")
                else:
                    self.log(f"No detailed checks found for {img_name} - JSON data: {'found' if json_data else 'not found'}")
                    
        except Exception as e:
            if self.debug:
                self.log(f"Error extracting detailed checks: {e}")
        
        # Process results for CSV
        mod_files_map = {}
        mod_status = ""
        csv_rows = []
        
        for line in results:
            parts = line.split(",")
            if len(parts) < 3:
                continue
            
            image_name, test_case, status = parts[0], parts[1], parts[2]
            
            if test_case == "HasModifiedFiles" and status == "FAILED":
                try:
                    with open(temp_log_file, "r") as lf:
                        log_content = lf.read()
                    files = re.findall(r'file=([^ ]+)', log_content)
                    mod_files_map[test_case] = ":".join(files)
                except Exception as e:
                    self.log(f"Error reading temp log: {e}")
                mod_status = "FAILED"
        
        # Build CSV rows
        for line in results:
            parts = line.split(",")
            if len(parts) < 3:
                continue
            
            image_name, test_case, status = parts[0], parts[1], parts[2].replace("ERROR", "NOT_APP")
            mod_files = mod_files_map.get(test_case, "") if mod_status == "FAILED" else ""
            csv_rows.append([image_name, tag, mod_files, test_case, status])
        
        if self.debug:
            self.log(f"Generated {len(csv_rows)} CSV rows for {img_name}")
        
        return results, csv_rows, detailed_checks

    def format_scan_output(self, results: List[str], img_name: str, repo_img_tag: str, verdict: str, elapsed: float) -> str:
        """Format the scan output for console display."""
        output = f"\nScanning image: {repo_img_tag}\n{'='*80}\n"
        output += f"{'Image Name':<36} {'Test Case':<26} {'Status':<10}\n"
        output += "-" * 79 + "\n"
        
        for line in results:
            parts = line.split(",")
            if len(parts) < 3:
                continue
            
            image_name, test_case, status = parts[0], parts[1], parts[2]
            
            if status == "FAILED":
                output += f"{image_name:<30} {test_case:<32} \033[1;31m{status:<12}\033[m\n"
            elif status == "PASSED":
                output += f"{image_name:<30} {test_case:<32} \033[1;32m{status:<12}\033[m\n"
            else:
                output += f"{image_name:<30} {test_case:<32} \033[1;33m{'NOT_APP':<12}\033[m\n"
        
        # Format verdict
        if verdict == "PASSED":
            verdict_colored = f"\033[1;32m{verdict}\033[m"
        else:
            verdict_colored = f"\033[1;31m{verdict}\033[m"
        
        output += f"Verdict: {verdict_colored}\n"
        output += f"Time elapsed: {elapsed:.3f} seconds\n"
        
        return output

    def scan_single_image(self, image: str) -> Dict[str, Any]:
        """Scan a single container image using preflight."""
        start_time = time.time()
        
        # Create temporary log file
        with tempfile.NamedTemporaryFile(delete=False, mode="w+", suffix=".log") as tmp_log:
            temp_log_file = tmp_log.name
        
        # Set environment variables
        old_logfile = os.environ.get("PFLT_LOGFILE")
        os.environ["PFLT_LOGFILE"] = temp_log_file
        os.environ["PFLT_JUNIT"] = "true"
        os.environ["PFLT_LOGLEVEL"] = "debug"
        
        try:
            # Log scanning mode
            mode = "single" if self.parallel == 1 else "parallel"
            self.log(f"Scanning image: {image} in {mode} mode")
            
            # Get image details
            image_info = self.get_image_details(image)
            if not image_info["success"]:
                raise Exception(image_info["error"])
            
            # Run preflight scan
            preflight_cmd = [
                "preflight", "check", "container", "--platform", "amd64", 
                image_info["inspect_url"]
            ]
            if self.auth_json:
                preflight_cmd.extend(["-d", self.auth_json])
            
            proc = subprocess.run(preflight_cmd, capture_output=True, text=True)
            exit_status = proc.returncode
            combined_output = proc.stdout + proc.stderr
            
            if self.debug:
                print(combined_output)
            
            # Wait for log file to be updated
            time.sleep(0.2)
            
            # Parse results
            results, csv_rows, detailed_checks = self.parse_preflight_output(
                combined_output, 
                image_info["img_name"], 
                image_info["tag"], 
                temp_log_file
            )
            
            # Get verdict
            verdict = "NOT_APP"
            try:
                with open(temp_log_file, "r") as lf:
                    log_content = lf.read()
                m_verdict = re.search(r'result:\s*(PASSED|FAILED|NOT_APP)', log_content)
                if m_verdict:
                    verdict = m_verdict.group(1).strip()
            except Exception as e:
                self.log(f"Error reading temporary log: {e}")
            
            if not verdict or verdict == "NOT_APP":
                m_verdict = re.search(r'Preflight result:\s*(PASSED|FAILED|NOT_APP)', combined_output)
                if m_verdict:
                    verdict = m_verdict.group(1).strip()
            
            # Format output
            elapsed = time.time() - start_time
            output = self.format_scan_output(
                results, 
                image_info["img_name"], 
                image_info["repo_img_tag"], 
                verdict, 
                elapsed
            )
            
            return {
                "error": exit_status != 0,
                "csv_rows": csv_rows,
                "detailed_checks": detailed_checks,
                "elapsed": elapsed,
                "console_output": output,
                "image": image
            }
            
        except Exception as e:
            err_msg = f"Error scanning {image}: {e}"
            self.log(err_msg)
            return {
                "error": True,
                "csv_rows": [],
                "detailed_checks": [],
                "elapsed": time.time() - start_time,
                "console_output": err_msg,
                "image": image
            }
        finally:
            # Cleanup
            if os.path.exists(temp_log_file):
                os.remove(temp_log_file)
            
            # Restore environment
            if old_logfile is not None:
                os.environ["PFLT_LOGFILE"] = old_logfile
            else:
                os.environ.pop("PFLT_LOGFILE", None)

    def scan_images_in_parallel(self) -> bool:
        """Scan multiple images in parallel using ThreadPoolExecutor."""
        total_start = time.time()
        all_scan_data = []
        all_detailed_checks = []
        combined_output = ""
        error_occurred = False
        count = 0

        with ThreadPoolExecutor(max_workers=self.parallel) as executor:
            future_to_image = {
                executor.submit(self.scan_single_image, image): image 
                for image in self.image_list
            }
            
            for future in as_completed(future_to_image):
                result = future.result()
                combined_output += result["console_output"]
                if result["error"]:
                    error_occurred = True
                all_scan_data.extend(result["csv_rows"])
                all_detailed_checks.extend(result.get("detailed_checks", []))
                count += 1

        # Calculate total time
        total_elapsed = time.time() - total_start
        
        # Add summary
        combined_output += "-" * 78 + "\n"
        combined_output += f"Total Images Scanned: {count}\n"
        combined_output += f"Total Scan Time: {time.strftime('%Hh:%Mm:%Ss', time.gmtime(total_elapsed))}\n"
        combined_output += "-" * 78 + "\n"

        # Write results to both XLSX and HTML reports
        if self.debug:
            self.log(f"Total scan data rows: {len(all_scan_data)}")
            self.log(f"Total detailed checks: {len(all_detailed_checks)}")
        
        if all_scan_data:
            self.write_results_to_reports(all_scan_data, all_detailed_checks)
        else:
            self.log("No scan data to write to reports - scan_data is empty")
        
        print(combined_output)
        return not error_occurred

    def run_prerequisite_checks(self) -> None:
        """Run all prerequisite checks."""
        print("\nChecking pre-requisite steps...")
        print("=" * 56)
        print(f"{'Pre-Requisites':<46} {'Status':<10}")
        print("-" * 57)
        
        self.check_required_tools()
        self.check_python_version()
        self.check_preflight_version()
        
        if self.fqdn:
            self.check_registry_connection()
        
        if self.api_token:
            self.check_registry_authentication()
        
        self.check_python_dependencies()
        print("=" * 55)

    def run(self) -> None:
        """Main execution method."""
        self.run_prerequisite_checks()
        self.build_image_list()
        
        # Backup existing results
        if self.file_exists(self.RESULT_XLSX):
            self.rename_file(self.RESULT_XLSX, self.RESULT_XLSX + "_saved")
        
        if self.file_exists(self.RESULT_HTML):
            self.rename_file(self.RESULT_HTML, self.RESULT_HTML + "_saved")
        
        # Scan images and write to both Excel and HTML reports
        if not self.scan_images_in_parallel():
            self.exit_with_error("Container image scanning failed.")

def main():
    """Main entry point."""
    scanner = PreflightScanner()
    scanner.run()

if __name__ == "__main__":
    main()
